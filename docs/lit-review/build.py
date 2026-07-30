#!/usr/bin/env python3
import csv, json, os, sys
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from annotations import ANN, NONARXIV, THEMES

SCRATCH = os.path.dirname(os.path.abspath(__file__))
OUTDIR = sys.argv[1]
os.makedirs(OUTDIR, exist_ok=True)
META = json.load(open(os.path.join(SCRATCH, "meta.json")))

FONT = "Arial"
NAVY = "1F3864"
HDR_FILL = PatternFill("solid", fgColor=NAVY)
BAND = PatternFill("solid", fgColor="EDF2F9")
YELLOW = PatternFill("solid", fgColor="FFF2CC")
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

THEME_READERS = {"T1": ("R01", "R02"), "T2": ("R03", "R04"), "T3": ("R05", "R06"),
                 "T4": ("R07", "R08"), "T5": ("R09", "R10")}
DUE = {1: date(2026, 8, 13), 2: date(2026, 9, 3), 3: date(2026, 9, 24)}

# ---------------------------------------------------------------- assemble rows
rows = []
for key, (theme, prio, concern, why, question) in ANN.items():
    if key.startswith("NONARXIV-"):
        title, authors, pub, venue, link = NONARXIV[key]
        alist = [a.strip() for a in authors.split(";")]
        source = "peer-reviewed"
    else:
        m = META[key]
        title = m["title"].replace("$^{+}$", "+").replace("$^2$", "^2").replace("$", "")
        alist = m["authors"]
        pub = m["published"]
        venue = m["comment"][:180] or "arXiv preprint"
        link = f"https://arxiv.org/abs/{key}"
        source = f"arXiv:{key}"
    rows.append(dict(key=key, theme=theme, prio=prio, concern=concern, why=why,
                     question=question, title=title, authors=alist, pub=pub,
                     venue=venue, link=link, source=source))

# stable order: theme, priority, date desc
rows.sort(key=lambda r: (r["theme"], r["prio"], r["pub"]), reverse=False)

# round-robin within theme so each reader gets a balanced priority mix
counters = {t: 0 for t in THEME_READERS}
for r in rows:
    a, b = THEME_READERS[r["theme"]]
    r["slot"] = a if counters[r["theme"]] % 2 == 0 else b
    counters[r["theme"]] += 1
for i, r in enumerate(rows, start=1):
    r["id"] = f"P{i:03d}"

# ---------------------------------------------------------------- workbook
wb = Workbook()

def style_header(ws, row, ncols, height=42):
    ws.row_dimensions[row].height = height
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
        cell.fill = HDR_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = BOX

# ============================================================ 1. README
ws = wb.active
ws.title = "README"
ws.sheet_view.showGridLines = False
readme = [
    ("Literature review tracker - prove-TLA", "h1"),
    ("Two concerns, 125 papers, 10 readers", "sub"),
    ("", ""),
    ("What this is", "h2"),
    ("A reading tracker for the two open questions on the prove-TLA project:", "p"),
    ("   (1) TRAINING - was the fine-tuning failure a method problem, a corpus problem, or a broken MoE adapter?", "p"),
    ("   (2) CONTAMINATION - how do we prove a holdout is clean, and what claims does the contaminated 206-spec corpus permit?", "p"),
    ("", ""),
    ("Every paper here was resolved against the live arXiv API or a publisher page. Titles, authors and", "p"),
    ("dates are as returned by the source, not typed from memory. The 'Why it matters' and 'Question this", "p"),
    ("should answer' columns are editorial - argue with them.", "p"),
    ("", ""),
    ("How to use it", "h2"),
    ("1. Open the Roster tab and replace 'Reader 01' ... 'Reader 10' in column B with real names.", "p"),
    ("   Every name in the Papers tab is a lookup on that column, so you only type each name once.", "p"),
    ("2. Each reader owns one half of a theme; the two readers on a theme cross-check each other's notes.", "p"),
    ("3. In the Papers tab, fill the yellow columns: Status, Key takeaway, Actionable here?, Cite in", "p"),
    ("   write-up?, Notes link, Reviewer notes. Leave everything left of Status alone.", "p"),
    ("4. Roster columns E-J recalculate from the Papers tab automatically.", "p"),
    ("", ""),
    ("Priorities and deadlines", "h2"),
    ("P1  must read      44 papers   due 2026-08-13   the papers a claim in the write-up will rest on", "p"),
    ("P2  should read    49 papers   due 2026-09-03   method detail, close comparators, standard references", "p"),
    ("P3  skim           32 papers   due 2026-09-24   adjacent or speculative; read the abstract, flag if it matters", "p"),
    ("", ""),
    ("Themes", "h2"),
]
for tid in sorted(THEMES):
    name, desc = THEMES[tid]
    a, b = THEME_READERS[tid]
    n = sum(1 for r in rows if r["theme"] == tid)
    readme.append((f"{tid}  {name}  -  {n} papers, readers {a} and {b}", "p"))
    readme.append((f"      {desc}", "small"))
readme += [
    ("", ""),
    ("Definition of done for one paper", "h2"),
    ("Status = Done means: the Key takeaway cell says something a colleague could act on without", "p"),
    ("reading the paper, and Actionable here? is answered Yes / No / Maybe with a reason in Reviewer notes.", "p"),
    ("A one-line summary of the abstract is not done.", "p"),
    ("", ""),
    ("Built 2026-07-30. Source metadata: arXiv API (export.arxiv.org), retrieved 2026-07-30.", "small"),
]
ws.column_dimensions["A"].width = 118
for i, (text, kind) in enumerate(readme, start=1):
    c = ws.cell(row=i, column=1, value=text)
    if kind == "h1":
        c.font = Font(name=FONT, bold=True, size=17, color=NAVY)
        ws.row_dimensions[i].height = 26
    elif kind == "sub":
        c.font = Font(name=FONT, size=11, color="7F7F7F")
    elif kind == "h2":
        c.font = Font(name=FONT, bold=True, size=12, color=NAVY)
        ws.row_dimensions[i].height = 20
    elif kind == "small":
        c.font = Font(name=FONT, size=9, color="595959")
    else:
        c.font = Font(name=FONT, size=10)
    c.alignment = Alignment(vertical="center")

# ============================================================ 2. Roster
ws = wb.create_sheet("Roster")
ws.sheet_view.showGridLines = False
ws["A1"] = "Roster - put real names in column B"
ws["A1"].font = Font(name=FONT, bold=True, size=14, color=NAVY)
ws["A2"] = ("Column B is the only place a name is typed. Papers column E looks it up by slot, so renaming here "
            "renames every assignment. Columns E-J are formulas - do not overwrite.")
ws["A2"].font = Font(name=FONT, size=9, color="595959")
ws.merge_cells("A2:I2")
ws["A2"].alignment = Alignment(wrap_text=True, vertical="center")
ws.row_dimensions[2].height = 26

hdr = ["Slot", "Name (EDIT ME)", "Theme", "Theme name", "Papers", "P1", "P2", "P3", "Done", "% done"]
for j, h in enumerate(hdr, start=1):
    ws.cell(row=4, column=j, value=h)
style_header(ws, 4, len(hdr), height=30)

slot_theme = {}
for t, (a, b) in THEME_READERS.items():
    slot_theme[a] = t
    slot_theme[b] = t
for i, slot in enumerate(sorted(slot_theme), start=5):
    t = slot_theme[slot]
    ws.cell(row=i, column=1, value=slot)
    nm = ws.cell(row=i, column=2, value=f"Reader {slot[1:]}")
    nm.fill = YELLOW
    ws.cell(row=i, column=3, value=t)
    ws.cell(row=i, column=4, value=THEMES[t][0])
    ws.cell(row=i, column=5, value=f'=COUNTIF(Papers!$D:$D,$A{i})')
    for k, p in enumerate([1, 2, 3]):
        ws.cell(row=i, column=6 + k,
                value=f'=COUNTIFS(Papers!$D:$D,$A{i},Papers!$C:$C,{p})')
    ws.cell(row=i, column=9, value=f'=COUNTIFS(Papers!$D:$D,$A{i},Papers!$Q:$Q,"Done")')
    ws.cell(row=i, column=10, value=f'=IFERROR($I{i}/$E{i},0)')
    ws.cell(row=i, column=10).number_format = "0%"
    for j in range(1, 11):
        c = ws.cell(row=i, column=j)
        c.font = Font(name=FONT, size=10, bold=(j == 2))
        c.border = BOX
        c.alignment = Alignment(vertical="center",
                                horizontal="left" if j in (2, 4) else "center")

tot = 15
ws.cell(row=tot, column=4, value="TOTAL").font = Font(name=FONT, bold=True, size=10)
for j in (5, 6, 7, 8, 9):
    c = ws.cell(row=tot, column=j, value=f'=SUM({get_column_letter(j)}5:{get_column_letter(j)}14)')
    c.font = Font(name=FONT, bold=True, size=10)
    c.border = BOX
    c.alignment = Alignment(horizontal="center")
c = ws.cell(row=tot, column=10, value="=IFERROR(I15/E15,0)")
c.font = Font(name=FONT, bold=True, size=10)
c.number_format = "0%"
c.border = BOX
c.alignment = Alignment(horizontal="center")

for col, w in zip("ABCDEFGHIJ", [8, 22, 8, 46, 9, 7, 7, 7, 8, 9]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = "A5"

# ============================================================ 3. Papers
ws = wb.create_sheet("Papers")
cols = [
    ("ID", 7), ("Theme", 7), ("Pri", 5), ("Reader slot", 9), ("Reader name", 18),
    ("Title", 60), ("First author", 20), ("All authors", 46), ("Year", 6),
    ("Published", 11), ("Venue / notes", 30), ("Source", 15), ("Link", 34),
    ("Concern", 13), ("Why it matters to prove-TLA", 62),
    ("Question this should answer for us", 58),
    ("Status", 14), ("Due", 11), ("Date completed", 13),
    ("Key takeaway (fill in)", 46), ("Actionable here?", 14),
    ("Cite in write-up?", 14), ("Notes link", 22), ("Reviewer notes", 46),
]
ws["A1"] = "Papers - one row per paper. Columns Q onward are yours to fill; everything left of them is source metadata."
ws["A1"].font = Font(name=FONT, bold=True, size=11, color=NAVY)
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
ws.row_dimensions[1].height = 20

for j, (h, w) in enumerate(cols, start=1):
    ws.cell(row=2, column=j, value=h)
    ws.column_dimensions[get_column_letter(j)].width = w
style_header(ws, 2, len(cols))

EDITABLE = set(range(17, 25))  # Q..X
for i, r in enumerate(rows, start=3):
    a = r["authors"]
    vals = [
        r["id"], r["theme"], r["prio"], r["slot"],
        f'=INDEX(Roster!$B$5:$B$14,MATCH($D{i},Roster!$A$5:$A$14,0))',
        r["title"], a[0] if a else "",
        "; ".join(a) if len(a) <= 8 else "; ".join(a[:8]) + f"; +{len(a)-8} more",
        int(r["pub"][:4]), r["pub"], r["venue"], r["source"], r["link"],
        r["concern"], r["why"], r["question"],
        "Not started", DUE[r["prio"]], None, None, None, None, None, None,
    ]
    for j, v in enumerate(vals, start=1):
        c = ws.cell(row=i, column=j, value=v)
        c.font = Font(name=FONT, size=9)
        c.border = BOX
        c.alignment = Alignment(vertical="top", wrap_text=(j in (6, 8, 11, 15, 16, 20, 24)))
        if j in EDITABLE:
            c.fill = YELLOW
        elif i % 2 == 1:
            c.fill = BAND
        if j in (2, 3, 4, 9, 14, 21, 22):
            c.alignment = Alignment(vertical="top", horizontal="center")
        if j == 3:
            c.font = Font(name=FONT, size=9, bold=True,
                          color={1: "C00000", 2: "BF8F00", 3: "595959"}[r["prio"]])
        if j == 13:
            c.font = Font(name=FONT, size=9, color="0563C1", underline="single")
            c.hyperlink = r["link"]
        if j == 18:
            c.number_format = "yyyy-mm-dd"
    ws.row_dimensions[i].height = 44

last = len(rows) + 2
dv_status = DataValidation(type="list", allow_blank=False, showDropDown=False,
    formula1='"Not started,Reading,Notes drafted,Discussed,Done,Dropped"')
dv_yn = DataValidation(type="list", allow_blank=True, showDropDown=False,
    formula1='"Yes,No,Maybe"')
dv_cite = DataValidation(type="list", allow_blank=True, showDropDown=False,
    formula1='"Core cite,Background cite,No"')
ws.add_data_validation(dv_status); dv_status.add(f"Q3:Q{last}")
ws.add_data_validation(dv_yn);     dv_yn.add(f"U3:U{last}")
ws.add_data_validation(dv_cite);   dv_cite.add(f"V3:V{last}")

ws.conditional_formatting.add(f"Q3:Q{last}", CellIsRule(
    operator="equal", formula=['"Done"'],
    fill=PatternFill("solid", fgColor="C6EFCE"), font=Font(name=FONT, size=9, color="006100")))
ws.conditional_formatting.add(f"Q3:Q{last}", CellIsRule(
    operator="equal", formula=['"Not started"'],
    font=Font(name=FONT, size=9, color="9C6500")))

ws.auto_filter.ref = f"A2:{get_column_letter(len(cols))}{last}"
ws.freeze_panes = "F3"

# ============================================================ 4. Themes
ws = wb.create_sheet("Themes")
ws.sheet_view.showGridLines = False
ws["A1"] = "Themes - what each pair of readers owns"
ws["A1"].font = Font(name=FONT, bold=True, size=14, color=NAVY)
hdr = ["Theme", "Name", "Readers", "Papers", "P1", "Maps to concern", "The question this theme owns"]
for j, h in enumerate(hdr, start=1):
    ws.cell(row=3, column=j, value=h)
style_header(ws, 3, len(hdr), height=28)
CONCERN_MAP = {"T1": "Both (state of the art)", "T2": "Both (method transfer)",
               "T3": "(1) Training", "T4": "(2) Contamination",
               "T5": "(1) + (2) bridge"}
for i, tid in enumerate(sorted(THEMES), start=4):
    name, desc = THEMES[tid]
    a, b = THEME_READERS[tid]
    ws.cell(row=i, column=1, value=tid)
    ws.cell(row=i, column=2, value=name)
    ws.cell(row=i, column=3, value=f"{a}, {b}")
    ws.cell(row=i, column=4, value=f'=COUNTIF(Papers!$B:$B,$A{i})')
    ws.cell(row=i, column=5, value=f'=COUNTIFS(Papers!$B:$B,$A{i},Papers!$C:$C,1)')
    ws.cell(row=i, column=6, value=CONCERN_MAP[tid])
    ws.cell(row=i, column=7, value=desc)
    for j in range(1, 8):
        c = ws.cell(row=i, column=j)
        c.font = Font(name=FONT, size=10)
        c.border = BOX
        c.alignment = Alignment(vertical="top", wrap_text=(j in (2, 7)),
                                horizontal="center" if j in (1, 4, 5) else "left")
    ws.row_dimensions[i].height = 58
for col, w in zip("ABCDEFG", [8, 44, 12, 9, 7, 22, 72]):
    ws.column_dimensions[col].width = w

xlsx = os.path.join(OUTDIR, "prove-TLA_literature_review.xlsx")
wb.save(xlsx)

# ============================================================ CSV mirror
csv_path = os.path.join(OUTDIR, "prove-TLA_literature_review.csv")
with open(csv_path, "w", newline="", encoding="utf-8") as f:
    w = csv.writer(f)
    w.writerow(["id", "theme", "theme_name", "priority", "reader_slot", "reader_name",
                "title", "first_author", "all_authors", "year", "published",
                "venue_notes", "source", "link", "concern", "why_it_matters",
                "question_for_us", "status", "due", "date_completed", "key_takeaway",
                "actionable", "cite_in_writeup", "notes_link", "reviewer_notes"])
    for r in rows:
        w.writerow([r["id"], r["theme"], THEMES[r["theme"]][0], r["prio"], r["slot"],
                    f"Reader {r['slot'][1:]}", r["title"], r["authors"][0] if r["authors"] else "",
                    "; ".join(r["authors"]), r["pub"][:4], r["pub"], r["venue"],
                    r["source"], r["link"], r["concern"], r["why"], r["question"],
                    "Not started", DUE[r["prio"]].isoformat(), "", "", "", "", "", ""])

from collections import Counter
print("papers:", len(rows))
print("by theme:", dict(sorted(Counter(r["theme"] for r in rows).items())))
print("by priority:", dict(sorted(Counter(r["prio"] for r in rows).items())))
print("by concern:", dict(Counter(r["concern"] for r in rows)))
print("per reader:", dict(sorted(Counter(r["slot"] for r in rows).items())))
print("wrote", xlsx)
print("wrote", csv_path)
